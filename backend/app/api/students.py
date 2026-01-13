"""
学生管理API路由模块
"""
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.services.base_service import BaseService
from app.api.deps import get_current_user, require_permission
from app.models.user import User
from app.schemas import (
    StudentInfo,
    StudentCreate,
    StudentUpdate,
    DeleteResponse,
    ImportResponse,
    ResponseBase
)

router = APIRouter(prefix="/students", tags=["学生管理"])


@router.get("", summary="获取学生列表")
async def get_students(
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页数量"),
    grade_id: int = Query(None, description="按年级筛选"),
    class_id: int = Query(None, description="按班级筛选"),
    keyword: str = Query(None, description="搜索关键词（姓名或学号）"),
    gender: str = Query(None, description="按性别筛选（M/F）"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    获取学生列表
    """
    base_service = BaseService(db)
    students, total = base_service.get_student_list(
        page=page,
        page_size=page_size,
        grade_id=grade_id,
        class_id=class_id,
        keyword=keyword,
        gender=gender
    )
    
    return {
        "items": [
            StudentInfo(
                id=s.id,
                class_id=s.class_id,
                student_no=s.student_no,
                name=s.name,
                gender=s.gender,
                class_name=s.class_.name if s.class_ else None,
                grade_name=s.class_.grade.name if s.class_ and s.class_.grade else None,
                created_at=s.created_at
            ) for s in students
        ],
        "total": total,
        "page": page,
        "page_size": page_size
    }


@router.post("", response_model=StudentInfo, summary="创建学生")
async def create_student(
    request: StudentCreate,
    current_user: User = Depends(require_permission("base_manage")),
    db: Session = Depends(get_db)
):
    """创建学生"""
    base_service = BaseService(db)
    student, error = base_service.create_student(
        class_id=request.class_id,
        student_no=request.student_no,
        name=request.name,
        gender=request.gender
    )
    
    if error:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=error)
    
    return StudentInfo(
        id=student.id,
        class_id=student.class_id,
        student_no=student.student_no,
        name=student.name,
        gender=student.gender,
        class_name=student.class_.name if student.class_ else None,
        grade_name=student.class_.grade.name if student.class_ and student.class_.grade else None
    )


# 固定路径路由必须在动态路径之前定义
@router.get("/template", summary="下载班级学生导入模板")
async def download_student_template(
    class_id: int = Query(..., description="班级ID"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """下载指定班级的学生导入Excel模板"""
    from fastapi.responses import StreamingResponse
    from urllib.parse import quote
    import io
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="服务器缺少openpyxl库"
        )
    
    # 查询班级信息
    from app.models.base import Class
    class_obj = db.query(Class).filter(Class.id == class_id).first()
    if not class_obj:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="班级不存在")
    
    grade_name = class_obj.grade.name if class_obj.grade else ""
    class_name = class_obj.name
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学生导入"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 表头：学号、姓名、性别
    headers = ["学号", "姓名", "性别"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # 示例数据
    sample_data = [
        ["2024001", "张三", "男"],
        ["2024002", "李四", "女"],
        ["2024003", "王五", "男"],
    ]
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
    
    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 8
    
    # 添加说明
    ws.cell(row=6, column=1, value=f"导入班级：{grade_name} {class_name}")
    ws.cell(row=7, column=1, value="说明：")
    ws.cell(row=8, column=1, value="1. 学号必须唯一")
    ws.cell(row=9, column=1, value="2. 性别填写：男 或 女")
    ws.cell(row=10, column=1, value="3. 请删除示例数据后再导入")
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 使用 RFC 5987 编码中文文件名
    filename = f"{grade_name}_{class_name}_学生导入模板.xlsx"
    encoded_filename = quote(filename)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        }
    )


@router.post("/import-by-class", response_model=ImportResponse, summary="按班级批量导入学生")
async def import_students_by_class(
    class_id: int = Query(..., description="班级ID"),
    file: UploadFile = File(..., description="Excel文件"),
    current_user: User = Depends(require_permission("base_manage")),
    db: Session = Depends(get_db)
):
    """按班级批量导入学生"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="只支持Excel文件格式(.xlsx, .xls)"
        )
    
    from app.models.base import Class
    class_obj = db.query(Class).filter(Class.id == class_id).first()
    if not class_obj:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="班级不存在")
    
    content = await file.read()
    base_service = BaseService(db)
    success, failed, errors = base_service.import_students_by_class(content, class_id)
    
    return ImportResponse(success=success, failed=failed, errors=errors)


@router.delete("/clear-all", response_model=ResponseBase, summary="清空所有学生")
async def clear_all_students(
    current_user: User = Depends(require_permission("base_manage")),
    db: Session = Depends(get_db)
):
    """清空所有学生"""
    from app.models.base import Student
    from sqlalchemy import text
    
    student_count = db.query(Student).delete()
    db.commit()
    
    # 重置自增ID（MySQL语法）
    db.execute(text("ALTER TABLE students AUTO_INCREMENT = 1"))
    db.commit()
    
    return ResponseBase(message=f"已清空 {student_count} 个学生")


# 动态路径路由放在最后
@router.get("/{student_id}", response_model=StudentInfo, summary="获取学生详情")
async def get_student(
    student_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """获取学生详情"""
    from app.models.base import Student
    student = db.query(Student).filter(Student.id == student_id).first()
    
    if not student:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="学生不存在")
    
    return StudentInfo(
        id=student.id,
        class_id=student.class_id,
        student_no=student.student_no,
        name=student.name,
        gender=student.gender,
        class_name=student.class_.name if student.class_ else None,
        grade_name=student.class_.grade.name if student.class_ and student.class_.grade else None
    )


@router.put("/{student_id}", response_model=StudentInfo, summary="更新学生信息")
async def update_student(
    student_id: int,
    request: StudentUpdate,
    current_user: User = Depends(require_permission("base_manage")),
    db: Session = Depends(get_db)
):
    """更新学生信息"""
    base_service = BaseService(db)
    student, error = base_service.update_student(
        student_id=student_id,
        class_id=request.class_id,
        student_no=request.student_no,
        name=request.name,
        gender=request.gender
    )
    
    if error:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=error)
    
    return StudentInfo(
        id=student.id,
        class_id=student.class_id,
        student_no=student.student_no,
        name=student.name,
        gender=student.gender,
        class_name=student.class_.name if student.class_ else None,
        grade_name=student.class_.grade.name if student.class_ and student.class_.grade else None
    )


@router.delete("/{student_id}", response_model=DeleteResponse, summary="删除学生")
async def delete_student(
    student_id: int,
    current_user: User = Depends(require_permission("base_manage")),
    db: Session = Depends(get_db)
):
    """删除学生"""
    base_service = BaseService(db)
    success, error, affected = base_service.delete_student(student_id)
    
    if not success:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=error)
    
    return DeleteResponse(message="删除成功", affected=affected if affected else None)
