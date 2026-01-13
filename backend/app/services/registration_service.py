"""
报名管理服务模块
实现报名创建、查重检测、限制校验、批量导入
"""
from typing import List, Optional, Tuple, Dict
from sqlalchemy.orm import Session
from sqlalchemy import func
from openpyxl import load_workbook
from io import BytesIO

from app.models.registration import Registration
from app.models.base import Student, Class
from app.models.event import Event, EventGroup


class RegistrationService:
    """报名管理服务类"""
    
    def __init__(self, db: Session):
        self.db = db
    
    def check_duplicate(self, student_id: int, event_id: int) -> Tuple[bool, Optional[Registration]]:
        """
        检查是否重复报名
        返回: (是否重复, 已存在的报名记录)
        """
        existing = self.db.query(Registration).filter(
            Registration.student_id == student_id,
            Registration.event_id == event_id
        ).first()
        return existing is not None, existing
    
    def check_class_limit(self, class_id: int, event_id: int) -> Tuple[bool, int, int]:
        """
        检查班级报名人数限制
        返回: (是否超限, 当前人数, 限制人数)
        """
        event = self.db.query(Event).filter(Event.id == event_id).first()
        if not event:
            return True, 0, 0
        
        # 统计该班级在该项目的报名人数
        current_count = self.db.query(Registration).join(Student).filter(
            Student.class_id == class_id,
            Registration.event_id == event_id
        ).count()
        
        return current_count >= event.max_per_class, current_count, event.max_per_class
    
    def check_student_limit(self, student_id: int) -> Tuple[bool, int, int]:
        """
        检查学生报名项目数限制
        返回: (是否超限, 当前项目数, 限制数)
        """
        # 获取任意一个项目的限制（假设所有项目限制相同）
        event = self.db.query(Event).first()
        max_per_student = event.max_per_student if event else 3
        
        current_count = self.db.query(Registration).filter(
            Registration.student_id == student_id
        ).count()
        
        return current_count >= max_per_student, current_count, max_per_student
    
    def create_registration(
        self,
        student_id: int,
        event_id: int,
        group_id: int = None,
        created_by: int = None
    ) -> Tuple[Optional[Registration], str]:
        """创建报名记录"""
        # 检查学生是否存在
        student = self.db.query(Student).filter(Student.id == student_id).first()
        if not student:
            return None, "学生不存在"
        
        # 检查项目是否存在
        event = self.db.query(Event).filter(Event.id == event_id).first()
        if not event:
            return None, "项目不存在"
        
        # 检查重复报名
        is_duplicate, existing = self.check_duplicate(student_id, event_id)
        if is_duplicate:
            return None, "该学生已报名此项目"
        
        # 检查班级人数限制
        is_over_class_limit, current, limit = self.check_class_limit(student.class_id, event_id)
        if is_over_class_limit:
            return None, f"该班级报名人数已达上限({limit}人)"
        
        # 检查个人项目数限制
        is_over_student_limit, current, limit = self.check_student_limit(student_id)
        if is_over_student_limit:
            return None, f"该学生报名项目数已达上限({limit}项)"
        
        # 检查组别
        if group_id:
            group = self.db.query(EventGroup).filter(EventGroup.id == group_id).first()
            if not group or group.event_id != event_id:
                return None, "组别不存在或不属于该项目"
            
            # 检查性别限制
            if group.gender != "A" and group.gender != student.gender:
                return None, "该组别不允许该性别参加"
            
            # 检查年级限制
            if group.grade_ids and student.class_.grade_id not in group.grade_ids:
                return None, "该组别不允许该年级参加"
        
        registration = Registration(
            student_id=student_id,
            event_id=event_id,
            group_id=group_id,
            created_by=created_by
        )
        self.db.add(registration)
        self.db.commit()
        self.db.refresh(registration)
        return registration, ""
    
    def delete_registration(self, registration_id: int) -> Tuple[bool, str]:
        """取消报名"""
        registration = self.db.query(Registration).filter(Registration.id == registration_id).first()
        if not registration:
            return False, "报名记录不存在"
        
        # 检查是否有成绩记录
        if registration.scores:
            return False, "该报名已有成绩记录，不能取消"
        
        self.db.delete(registration)
        self.db.commit()
        return True, ""
    
    def clear_all_registrations(self) -> int:
        """清空所有报名记录和相关学生数据，并重置自增ID"""
        from app.models.base import Grade
        from sqlalchemy import text
        
        # 先删除所有报名记录
        reg_count = self.db.query(Registration).delete()
        
        # 删除所有学生
        self.db.query(Student).delete()
        
        # 删除所有班级
        self.db.query(Class).delete()
        
        # 删除所有年级
        self.db.query(Grade).delete()
        
        self.db.commit()
        
        # 重置自增ID（MySQL语法）
        self.db.execute(text("ALTER TABLE registrations AUTO_INCREMENT = 1"))
        self.db.execute(text("ALTER TABLE students AUTO_INCREMENT = 1"))
        self.db.execute(text("ALTER TABLE classes AUTO_INCREMENT = 1"))
        self.db.execute(text("ALTER TABLE grades AUTO_INCREMENT = 1"))
        self.db.commit()
        
        return reg_count
    
    def get_registration_list(
        self,
        page: int = 1,
        page_size: int = 20,
        event_id: int = None,
        group_id: int = None,
        group_name: str = None,
        class_id: int = None,
        grade_id: int = None,
        student_id: int = None,
        student_name: str = None
    ) -> Tuple[List[Registration], int]:
        """获取报名列表"""
        query = self.db.query(Registration).join(Student).join(Class)
        
        if event_id:
            query = query.filter(Registration.event_id == event_id)
        if group_id:
            query = query.filter(Registration.group_id == group_id)
        if group_name:
            # 按组别名称筛选，需要join EventGroup
            query = query.join(EventGroup, Registration.group_id == EventGroup.id)
            query = query.filter(EventGroup.name == group_name)
        if class_id:
            query = query.filter(Student.class_id == class_id)
        if grade_id:
            query = query.filter(Class.grade_id == grade_id)
        if student_id:
            query = query.filter(Registration.student_id == student_id)
        if student_name:
            query = query.filter(Student.name.like(f"%{student_name}%"))
        
        total = query.count()
        registrations = query.offset((page - 1) * page_size).limit(page_size).all()
        
        return registrations, total
    
    def update_lane_no(self, registration_id: int, lane_no: int) -> Tuple[bool, str]:
        """更新道次/序号"""
        registration = self.db.query(Registration).filter(Registration.id == registration_id).first()
        if not registration:
            return False, "报名记录不存在"
        
        registration.lane_no = lane_no
        self.db.commit()
        return True, ""
    
    def import_registrations(self, file_content: bytes, created_by: int = None, filename: str = None) -> Tuple[int, int, List[str]]:
        """
        批量导入报名 - 智能识别多种Excel格式
        
        导入时会自动创建学生信息（如果不存在），不需要预先导入学生数据
        报名时间使用当前导入时间
        
        支持的格式：
        1. 班级报名表格式（导出的格式）:
           - 组别标题行: 【项目名称 - 组别名称】（每班限报X人）
           - 表头行: 序号, 学号, 姓名, 性别, 项目名称, 组别名称
           - 数据行: 1, 0003, 张三, 男, 50米, 男子组
        
        返回: (成功数, 失败数, 错误列表)
        """
        from app.models.base import Grade
        from datetime import datetime
        
        success_count = 0
        fail_count = 0
        errors = []
        file_prefix = f"[{filename}] " if filename else ""
        
        # 记录导入时间
        import_time = datetime.now()
        
        try:
            wb = load_workbook(BytesIO(file_content))
            
            # 遍历所有工作表
            for ws in wb.worksheets:
                sheet_name = ws.title
                current_event_name = None
                current_group_name = None
                column_mapping = None
                
                # 从工作表名称解析班级信息（格式：年级名-班级名）
                class_info = self._parse_class_from_sheet_name(sheet_name)
                
                for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
                    if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                        continue
                    
                    first_cell = str(row[0]).strip() if row[0] is not None else ''
                    
                    # 跳过班级标题行，但尝试从中解析班级信息
                    if '运动会报名表' in first_cell or '报名表' in first_cell:
                        parsed = self._parse_class_from_title(first_cell)
                        if parsed:
                            class_info = parsed
                        continue
                    
                    # 检查是否是组别标题行
                    if first_cell.startswith('【') and ' - ' in first_cell:
                        try:
                            title_part = first_cell.split('】')[0].replace('【', '')
                            parts = title_part.split(' - ')
                            current_event_name = parts[0].strip()
                            current_group_name = parts[1].strip() if len(parts) > 1 else None
                            if current_group_name == '默认组':
                                current_group_name = None
                        except:
                            pass
                        # 不重置column_mapping，保持之前的映射
                        continue
                    
                    # 检查是否是表头行
                    row_values = [str(cell).strip() if cell else '' for cell in row]
                    if self._is_header_row(row_values):
                        column_mapping = self._detect_column_mapping(row_values)
                        continue
                    
                    # 使用默认映射
                    if column_mapping is None:
                        column_mapping = {
                            'student_no': 1,
                            'student_name': 2,
                            'gender': 3,
                            'event_name': 4,
                            'group_name': 5
                        }
                    
                    try:
                        # 解析数据行
                        student_no = self._get_cell_value(row, column_mapping.get('student_no'))
                        student_name = self._get_cell_value(row, column_mapping.get('student_name'))
                        gender_str = self._get_cell_value(row, column_mapping.get('gender'))
                        event_name = self._get_cell_value(row, column_mapping.get('event_name'))
                        group_name = self._get_cell_value(row, column_mapping.get('group_name'))
                        
                        # 使用当前组别标题中的信息作为默认值
                        if not event_name:
                            event_name = current_event_name
                        if not group_name:
                            group_name = current_group_name
                        
                        # 跳过无效数据（必须有学号和项目名称）
                        if not student_no or not event_name:
                            continue
                        
                        # 跳过空姓名的行（可能是预留的空行）
                        if not student_name:
                            continue
                        
                        # 处理组别名称
                        if group_name in ['默认组', '-', '']:
                            group_name = None
                        
                        # 转换性别
                        gender = 'M'
                        if gender_str in ['女', 'F', 'f', 'female']:
                            gender = 'F'
                        
                        # 查找项目（先查项目，避免创建无用的学生数据）
                        event = self.db.query(Event).filter(Event.name == event_name).first()
                        if not event:
                            errors.append(f"{file_prefix}[{sheet_name}] 第{row_idx}行: 项目'{event_name}'不存在")
                            fail_count += 1
                            continue
                        
                        # 查找或创建班级
                        class_obj = self._get_or_create_class(class_info)
                        if not class_obj:
                            errors.append(f"{file_prefix}[{sheet_name}] 第{row_idx}行: 无法确定班级信息")
                            fail_count += 1
                            continue
                        
                        # 查找或创建学生
                        student = self._get_or_create_student(
                            student_no=student_no,
                            student_name=student_name,
                            gender=gender,
                            class_id=class_obj.id
                        )
                        
                        # 查找组别
                        group_id = None
                        if group_name:
                            group = self.db.query(EventGroup).filter(
                                EventGroup.event_id == event.id,
                                EventGroup.name == group_name
                            ).first()
                            if group:
                                group_id = group.id
                        
                        # 检查是否重复报名（在数据库层面检查）
                        existing = self.db.query(Registration).filter(
                            Registration.student_id == student.id,
                            Registration.event_id == event.id
                        ).first()
                        
                        if existing:
                            # 重复报名，跳过
                            continue
                        
                        # 创建报名 - 使用 savepoint 来隔离每条记录的事务
                        try:
                            # 创建 savepoint
                            savepoint = self.db.begin_nested()
                            
                            registration = Registration(
                                student_id=student.id,
                                event_id=event.id,
                                group_id=group_id,
                                created_by=created_by,
                                created_at=import_time,
                                updated_at=import_time
                            )
                            self.db.add(registration)
                            self.db.flush()
                            success_count += 1
                        except Exception as reg_error:
                            # 回滚到 savepoint，不影响其他记录
                            self.db.rollback()
                            errors.append(f"{file_prefix}[{sheet_name}] 第{row_idx}行: {str(reg_error)}")
                            fail_count += 1
                            
                    except Exception as e:
                        errors.append(f"{file_prefix}[{sheet_name}] 第{row_idx}行: {str(e)}")
                        fail_count += 1
            
            self.db.commit()
            
        except Exception as e:
            self.db.rollback()
            errors.append(f"{file_prefix}文件解析错误: {str(e)}")
        
        return success_count, fail_count, errors
    
    def _parse_class_from_sheet_name(self, sheet_name: str) -> Dict:
        """从工作表名称解析班级信息（格式：年级名-班级名）"""
        if '-' in sheet_name:
            parts = sheet_name.split('-')
            return {
                'grade_name': parts[0].strip(),
                'class_name': parts[1].strip()
            }
        return None
    
    def _parse_class_from_title(self, title: str) -> Dict:
        """从标题行解析班级信息（格式：XX年级 XX班 运动会报名表）"""
        import re
        # 匹配 "一年级 1班" 或 "一年级1班" 格式
        match = re.search(r'([一二三四五六七八九十\d]+年级)\s*(\d+班)', title)
        if match:
            return {
                'grade_name': match.group(1),
                'class_name': match.group(2)
            }
        return None
    
    def _get_or_create_class(self, class_info: Dict) -> Class:
        """获取或创建班级"""
        from app.models.base import Grade
        
        if not class_info:
            return None
        
        grade_name = class_info.get('grade_name')
        class_name = class_info.get('class_name')
        
        if not grade_name or not class_name:
            return None
        
        # 查找或创建年级
        grade = self.db.query(Grade).filter(Grade.name == grade_name).first()
        if not grade:
            grade = Grade(name=grade_name, sort_order=0)
            self.db.add(grade)
            self.db.flush()
        
        # 查找或创建班级
        class_obj = self.db.query(Class).filter(
            Class.grade_id == grade.id,
            Class.name == class_name
        ).first()
        if not class_obj:
            class_obj = Class(name=class_name, grade_id=grade.id)
            self.db.add(class_obj)
            self.db.flush()
        
        return class_obj
    
    def _get_or_create_student(self, student_no: str, student_name: str, gender: str, class_id: int) -> Student:
        """获取或创建学生"""
        student = self.db.query(Student).filter(Student.student_no == student_no).first()
        if not student:
            student = Student(
                student_no=student_no,
                name=student_name,
                gender=gender,
                class_id=class_id
            )
            self.db.add(student)
            self.db.flush()
        return student
    
    def _detect_column_mapping(self, header_row: List[str]) -> Dict[str, int]:
        """
        根据表头行检测列映射
        """
        mapping = {}
        
        for idx, cell in enumerate(header_row):
            cell_lower = cell.lower() if cell else ''
            
            # 学号列
            if '学号' in cell or cell_lower == 'student_no':
                mapping['student_no'] = idx
            # 姓名列
            elif '姓名' in cell or cell_lower == 'name':
                mapping['student_name'] = idx
            # 性别列
            elif '性别' in cell or cell_lower == 'gender':
                mapping['gender'] = idx
            # 项目名称列
            elif '项目' in cell or cell_lower in ['event', 'event_name']:
                mapping['event_name'] = idx
            # 组别名称列
            elif '组别' in cell or cell_lower in ['group', 'group_name']:
                mapping['group_name'] = idx
        
        # 如果没有找到学号列，尝试其他方式
        if 'student_no' not in mapping:
            for idx, cell in enumerate(header_row):
                if cell in ['编号', 'ID', '学生编号']:
                    mapping['student_no'] = idx
                    break
        
        return mapping
    
    def _is_header_row(self, row_values: List[str]) -> bool:
        """
        判断是否是表头行
        """
        header_keywords = ['学号', '项目', '组别', '姓名', '性别', '序号', 'ID', '编号']
        match_count = sum(1 for cell in row_values if any(kw in cell for kw in header_keywords))
        return match_count >= 2
    
    def _get_cell_value(self, row: tuple, col_idx: int) -> str:
        """
        安全获取单元格值
        """
        if col_idx is None or col_idx >= len(row):
            return ''
        value = row[col_idx]
        if value is None:
            return ''
        return str(value).strip()
