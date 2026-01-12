"""
基础信息服务模块
实现年级/班级/学生的CRUD操作、关联数据检查、批量导入
"""
from typing import List, Optional, Tuple, Dict, Any
from sqlalchemy.orm import Session
from openpyxl import load_workbook
from io import BytesIO

from app.models.base import Grade, Class, Student
from app.models.registration import Registration


class BaseService:
    """基础信息服务类"""
    
    def __init__(self, db: Session):
        self.db = db
    
    # ========== 年级管理 ==========
    
    def create_grade(self, name: str, sort_order: int = 0) -> Tuple[Optional[Grade], str]:
        """创建年级"""
        existing = self.db.query(Grade).filter(Grade.name == name).first()
        if existing:
            return None, "年级名称已存在"
        
        grade = Grade(name=name, sort_order=sort_order)
        self.db.add(grade)
        self.db.commit()
        self.db.refresh(grade)
        return grade, ""
    
    def update_grade(self, grade_id: int, name: str = None, sort_order: int = None) -> Tuple[Optional[Grade], str]:
        """更新年级"""
        grade = self.db.query(Grade).filter(Grade.id == grade_id).first()
        if not grade:
            return None, "年级不存在"
        
        if name and name != grade.name:
            existing = self.db.query(Grade).filter(Grade.name == name).first()
            if existing:
                return None, "年级名称已存在"
            grade.name = name
        
        if sort_order is not None:
            grade.sort_order = sort_order
        
        self.db.commit()
        self.db.refresh(grade)
        return grade, ""
    
    def delete_grade(self, grade_id: int) -> Tuple[bool, str, Dict[str, int]]:
        """删除年级，返回关联数据统计"""
        grade = self.db.query(Grade).filter(Grade.id == grade_id).first()
        if not grade:
            return False, "年级不存在", {}
        
        # 检查关联数据
        class_count = len(grade.classes)
        student_count = sum(len(c.students) for c in grade.classes)
        
        if class_count > 0:
            return False, f"该年级下有{class_count}个班级，{student_count}名学生，请先删除关联数据", {
                "classes": class_count,
                "students": student_count
            }
        
        self.db.delete(grade)
        self.db.commit()
        return True, "", {}
    
    def get_grade_list(self) -> List[Grade]:
        """获取年级列表"""
        return self.db.query(Grade).order_by(Grade.sort_order, Grade.id).all()
    
    # ========== 班级管理 ==========
    
    def create_class(self, grade_id: int, name: str) -> Tuple[Optional[Class], str]:
        """创建班级"""
        grade = self.db.query(Grade).filter(Grade.id == grade_id).first()
        if not grade:
            return None, "年级不存在"
        
        existing = self.db.query(Class).filter(
            Class.grade_id == grade_id, 
            Class.name == name
        ).first()
        if existing:
            return None, "该年级下已存在同名班级"
        
        class_ = Class(grade_id=grade_id, name=name)
        self.db.add(class_)
        self.db.commit()
        self.db.refresh(class_)
        return class_, ""
    
    def update_class(self, class_id: int, name: str = None, grade_id: int = None) -> Tuple[Optional[Class], str]:
        """更新班级"""
        class_ = self.db.query(Class).filter(Class.id == class_id).first()
        if not class_:
            return None, "班级不存在"
        
        target_grade_id = grade_id if grade_id else class_.grade_id
        target_name = name if name else class_.name
        
        if name or grade_id:
            existing = self.db.query(Class).filter(
                Class.grade_id == target_grade_id,
                Class.name == target_name,
                Class.id != class_id
            ).first()
            if existing:
                return None, "该年级下已存在同名班级"
        
        if grade_id:
            grade = self.db.query(Grade).filter(Grade.id == grade_id).first()
            if not grade:
                return None, "目标年级不存在"
            class_.grade_id = grade_id
        
        if name:
            class_.name = name
        
        self.db.commit()
        self.db.refresh(class_)
        return class_, ""
    
    def delete_class(self, class_id: int) -> Tuple[bool, str, Dict[str, int]]:
        """删除班级"""
        class_ = self.db.query(Class).filter(Class.id == class_id).first()
        if not class_:
            return False, "班级不存在", {}
        
        student_count = len(class_.students)
        if student_count > 0:
            return False, f"该班级下有{student_count}名学生，请先删除学生", {
                "students": student_count
            }
        
        self.db.delete(class_)
        self.db.commit()
        return True, "", {}
    
    def get_class_list(self, grade_id: int = None) -> List[Class]:
        """获取班级列表"""
        query = self.db.query(Class)
        if grade_id:
            query = query.filter(Class.grade_id == grade_id)
        return query.order_by(Class.grade_id, Class.id).all()
    
    # ========== 学生管理 ==========
    
    def create_student(
        self, 
        class_id: int, 
        student_no: str, 
        name: str, 
        gender: str
    ) -> Tuple[Optional[Student], str]:
        """创建学生"""
        class_ = self.db.query(Class).filter(Class.id == class_id).first()
        if not class_:
            return None, "班级不存在"
        
        existing = self.db.query(Student).filter(Student.student_no == student_no).first()
        if existing:
            return None, "学号已存在"
        
        if gender not in ("M", "F"):
            return None, "性别必须是M(男)或F(女)"
        
        student = Student(
            class_id=class_id,
            student_no=student_no,
            name=name,
            gender=gender
        )
        self.db.add(student)
        self.db.commit()
        self.db.refresh(student)
        return student, ""
    
    def update_student(
        self,
        student_id: int,
        class_id: int = None,
        student_no: str = None,
        name: str = None,
        gender: str = None
    ) -> Tuple[Optional[Student], str]:
        """更新学生"""
        student = self.db.query(Student).filter(Student.id == student_id).first()
        if not student:
            return None, "学生不存在"
        
        if class_id:
            class_ = self.db.query(Class).filter(Class.id == class_id).first()
            if not class_:
                return None, "班级不存在"
            student.class_id = class_id
        
        if student_no and student_no != student.student_no:
            existing = self.db.query(Student).filter(Student.student_no == student_no).first()
            if existing:
                return None, "学号已存在"
            student.student_no = student_no
        
        if name:
            student.name = name
        
        if gender:
            if gender not in ("M", "F"):
                return None, "性别必须是M(男)或F(女)"
            student.gender = gender
        
        self.db.commit()
        self.db.refresh(student)
        return student, ""
    
    def delete_student(self, student_id: int) -> Tuple[bool, str, Dict[str, int]]:
        """删除学生"""
        student = self.db.query(Student).filter(Student.id == student_id).first()
        if not student:
            return False, "学生不存在", {}
        
        # 检查报名记录
        reg_count = self.db.query(Registration).filter(Registration.student_id == student_id).count()
        if reg_count > 0:
            return False, f"该学生有{reg_count}条报名记录，请先取消报名", {
                "registrations": reg_count
            }
        
        self.db.delete(student)
        self.db.commit()
        return True, "", {}
    
    def get_student_list(
        self,
        page: int = 1,
        page_size: int = 20,
        grade_id: int = None,
        class_id: int = None,
        keyword: str = None,
        gender: str = None
    ) -> Tuple[List[Student], int]:
        """获取学生列表"""
        query = self.db.query(Student).join(Class)
        
        if grade_id:
            query = query.filter(Class.grade_id == grade_id)
        if class_id:
            query = query.filter(Student.class_id == class_id)
        if keyword:
            query = query.filter(
                (Student.name.contains(keyword)) |
                (Student.student_no.contains(keyword))
            )
        if gender:
            query = query.filter(Student.gender == gender)
        
        total = query.count()
        students = query.offset((page - 1) * page_size).limit(page_size).all()
        
        return students, total
    
    def import_students(self, file_content: bytes) -> Tuple[int, int, List[str]]:
        """
        批量导入学生
        Excel格式: 学号, 姓名, 性别(男/女), 年级, 班级
        返回: (成功数, 失败数, 错误列表)
        """
        success_count = 0
        fail_count = 0
        errors = []
        
        try:
            wb = load_workbook(BytesIO(file_content))
            ws = wb.active
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:  # 跳过空行
                    continue
                
                try:
                    student_no = str(row[0]).strip()
                    name = str(row[1]).strip()
                    gender_str = str(row[2]).strip()
                    grade_name = str(row[3]).strip()
                    class_name = str(row[4]).strip()
                    
                    # 转换性别
                    gender = "M" if gender_str == "男" else "F" if gender_str == "女" else None
                    if not gender:
                        errors.append(f"第{row_idx}行: 性别格式错误，应为'男'或'女'")
                        fail_count += 1
                        continue
                    
                    # 查找年级
                    grade = self.db.query(Grade).filter(Grade.name == grade_name).first()
                    if not grade:
                        errors.append(f"第{row_idx}行: 年级'{grade_name}'不存在")
                        fail_count += 1
                        continue
                    
                    # 查找班级
                    class_ = self.db.query(Class).filter(
                        Class.grade_id == grade.id,
                        Class.name == class_name
                    ).first()
                    if not class_:
                        errors.append(f"第{row_idx}行: 班级'{class_name}'不存在")
                        fail_count += 1
                        continue
                    
                    # 检查学号是否已存在
                    existing = self.db.query(Student).filter(Student.student_no == student_no).first()
                    if existing:
                        errors.append(f"第{row_idx}行: 学号'{student_no}'已存在")
                        fail_count += 1
                        continue
                    
                    # 创建学生
                    student = Student(
                        class_id=class_.id,
                        student_no=student_no,
                        name=name,
                        gender=gender
                    )
                    self.db.add(student)
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"第{row_idx}行: {str(e)}")
                    fail_count += 1
            
            self.db.commit()
            
        except Exception as e:
            errors.append(f"文件解析错误: {str(e)}")
        
        return success_count, fail_count, errors

    def import_students_by_class(self, file_content: bytes, class_id: int) -> Tuple[int, int, List[str]]:
        """
        按班级批量导入学生
        Excel格式: 学号, 姓名, 性别(男/女)
        返回: (成功数, 失败数, 错误列表)
        """
        success_count = 0
        fail_count = 0
        errors = []
        
        try:
            wb = load_workbook(BytesIO(file_content))
            ws = wb.active
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:  # 跳过空行
                    continue
                
                try:
                    student_no = str(row[0]).strip()
                    name = str(row[1]).strip()
                    gender_str = str(row[2]).strip()
                    
                    # 转换性别
                    gender = "M" if gender_str == "男" else "F" if gender_str == "女" else None
                    if not gender:
                        errors.append(f"第{row_idx}行: 性别格式错误，应为'男'或'女'")
                        fail_count += 1
                        continue
                    
                    # 检查学号是否已存在
                    existing = self.db.query(Student).filter(Student.student_no == student_no).first()
                    if existing:
                        errors.append(f"第{row_idx}行: 学号'{student_no}'已存在")
                        fail_count += 1
                        continue
                    
                    # 创建学生
                    student = Student(
                        class_id=class_id,
                        student_no=student_no,
                        name=name,
                        gender=gender
                    )
                    self.db.add(student)
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"第{row_idx}行: {str(e)}")
                    fail_count += 1
            
            self.db.commit()
            
        except Exception as e:
            errors.append(f"文件解析错误: {str(e)}")
        
        return success_count, fail_count, errors
