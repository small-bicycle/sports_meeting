"""
数据导出服务模块
实现报名表导出、成绩表导出、排名表导出、参赛表格生成
"""
from typing import List, Dict, Optional
from io import BytesIO
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from app.models.registration import Registration
from app.models.score import Score
from app.models.base import Student, Class, Grade
from app.models.event import Event
from app.services.statistics_service import StatisticsService


class ExportService:
    """数据导出服务类"""
    
    def __init__(self, db: Session):
        self.db = db
        self.stats_service = StatisticsService(db)
    
    def _create_workbook(self) -> Workbook:
        """创建工作簿"""
        return Workbook()
    
    def _style_header(self, ws, row: int = 1):
        """设置表头样式"""
        header_font = Font(bold=True)
        for cell in ws[row]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
    
    def export_registration_form(
        self,
        event_id: int = None,
        class_id: int = None,
        grade_id: int = None
    ) -> bytes:
        """
        导出报名表 - 每个班级一个独立的Excel文件，打包成ZIP
        
        方便班主任单独填写各班级的报名信息
        """
        import zipfile
        from app.models.event import EventGroup
        
        # 查询所有班级（按年级和班级名排序）
        classes_query = (
            self.db.query(Class)
            .join(Grade, Class.grade_id == Grade.id)
            .order_by(Grade.sort_order, Class.name)
        )
        
        if class_id:
            classes_query = classes_query.filter(Class.id == class_id)
        if grade_id:
            classes_query = classes_query.filter(Class.grade_id == grade_id)
        
        classes = classes_query.all()
        
        # 查询所有项目及其组别
        events_query = self.db.query(Event).order_by(Event.name)
        if event_id:
            events_query = events_query.filter(Event.id == event_id)
        events = events_query.all()
        
        # 创建ZIP文件
        zip_buffer = BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # 为每个班级创建一个独立的Excel文件
            for class_ in classes:
                excel_content = self._create_class_registration_excel(
                    class_=class_,
                    events=events
                )
                
                # 文件名：年级名-班级名.xlsx
                filename = f"{class_.grade.name}-{class_.name}.xlsx"
                zip_file.writestr(filename, excel_content)
        
        zip_buffer.seek(0)
        return zip_buffer.getvalue()
    
    def _create_class_registration_excel(self, class_, events) -> bytes:
        """
        为单个班级创建报名表Excel
        """
        from app.models.event import EventGroup
        
        wb = self._create_workbook()
        ws = wb.active
        ws.title = "报名表"
        
        # 定义样式
        header_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        event_title_font = Font(bold=True, size=12)
        event_title_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        current_row = 1
        
        # 添加班级标题
        title_text = f"{class_.grade.name} {class_.name} 运动会报名表"
        ws.cell(row=current_row, column=1, value=title_text)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        title_cell = ws.cell(row=current_row, column=1)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = center_align
        current_row += 2
        
        # 遍历每个项目
        for event in events:
            # 获取该项目的组别
            groups = self.db.query(EventGroup).filter(EventGroup.event_id == event.id).all()
            
            # 如果项目没有组别，创建一个默认组
            if not groups:
                group_list = [{"name": "默认组", "id": None}]
            else:
                group_list = [{"name": g.name, "id": g.id} for g in groups]
            
            # 为每个组别创建一个区域
            for group_info in group_list:
                group_name = group_info["name"]
                group_id = group_info["id"]
                
                # 添加项目-组别标题行
                title_text = f"【{event.name} - {group_name}】（每班限报{event.max_per_class}人）"
                ws.cell(row=current_row, column=1, value=title_text)
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
                
                title_cell = ws.cell(row=current_row, column=1)
                title_cell.font = event_title_font
                title_cell.fill = event_title_fill
                title_cell.alignment = center_align
                
                current_row += 1
                
                # 添加表头
                headers = ["序号", "学号", "姓名", "性别", "项目名称", "组别名称"]
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=header)
                    cell.font = header_font
                    cell.alignment = center_align
                    cell.border = thin_border
                
                current_row += 1
                
                # 查询该班级在该项目-组别下的报名记录
                reg_query = (
                    self.db.query(Registration)
                    .join(Student, Registration.student_id == Student.id)
                    .filter(Student.class_id == class_.id)
                    .filter(Registration.event_id == event.id)
                )
                
                if group_id:
                    reg_query = reg_query.filter(Registration.group_id == group_id)
                else:
                    reg_query = reg_query.filter(Registration.group_id == None)
                
                registrations = reg_query.order_by(Student.student_no).all()
                
                # 添加已报名的学生数据
                for seq_no, reg in enumerate(registrations, 1):
                    gender_display = "男" if reg.student.gender == "M" else "女"
                    row_data = [
                        seq_no,
                        reg.student.student_no,
                        reg.student.name,
                        gender_display,
                        event.name,
                        group_name
                    ]
                    
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.alignment = center_align
                        cell.border = thin_border
                    
                    current_row += 1
                
                # 添加空行供班主任填写（预留到max_per_class行）
                empty_rows = max(0, event.max_per_class - len(registrations))
                start_seq = len(registrations) + 1
                for i in range(empty_rows):
                    row_data = [
                        start_seq + i,
                        "",  # 学号
                        "",  # 姓名
                        "",  # 性别
                        event.name,
                        group_name
                    ]
                    
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.alignment = center_align
                        cell.border = thin_border
                    
                    current_row += 1
                
                # 组别之间添加空行
                current_row += 1
        
        # 设置列宽
        column_widths = [8, 15, 12, 8, 15, 15]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def export_score_sheet(
        self,
        event_id: int = None,
        class_id: int = None,
        grade_id: int = None
    ) -> bytes:
        """导出成绩表"""
        wb = self._create_workbook()
        ws = wb.active
        ws.title = "成绩表"
        
        headers = ["序号", "学号", "姓名", "班级", "年级", "项目", "成绩", "轮次", "排名", "得分"]
        ws.append(headers)
        self._style_header(ws)
        
        query = self.db.query(Score).join(Registration).join(Student).join(Class).join(Grade).filter(
            Score.is_valid == True
        )
        
        if event_id:
            query = query.filter(Registration.event_id == event_id)
        if class_id:
            query = query.filter(Student.class_id == class_id)
        if grade_id:
            query = query.filter(Class.grade_id == grade_id)
        
        scores = query.all()
        
        for idx, score in enumerate(scores, 1):
            reg = score.registration
            ws.append([
                idx,
                reg.student.student_no,
                reg.student.name,
                reg.student.class_.name,
                reg.student.class_.grade.name,
                reg.event.name,
                float(score.value),
                "预赛" if score.round == "preliminary" else "决赛",
                score.rank or "",
                score.points or 0
            ])
        
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 12
        
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def export_ranking_sheet(self, type: str = "event", event_id: int = None) -> bytes:
        """导出排名表"""
        wb = self._create_workbook()
        ws = wb.active
        
        if type == "event":
            ws.title = "项目排名"
            headers = ["排名", "学号", "姓名", "班级", "年级", "成绩", "得分"]
            ws.append(headers)
            self._style_header(ws)
            
            if event_id:
                rankings = self.stats_service.get_event_ranking(event_id)
                for r in rankings:
                    ws.append([
                        r["rank"],
                        r["student"]["student_no"],
                        r["student"]["name"],
                        r["student"]["class_name"],
                        r["student"]["grade_name"],
                        r["score"]["value"],
                        r["points"]
                    ])
        
        elif type == "class":
            ws.title = "班级总分"
            headers = ["排名", "班级", "年级", "总分", "金牌", "银牌", "铜牌"]
            ws.append(headers)
            self._style_header(ws)
            
            rankings = self.stats_service.get_class_total()
            for r in rankings:
                ws.append([
                    r["rank"],
                    r["class"]["name"],
                    r["class"]["grade_name"],
                    r["total_score"],
                    r["gold"],
                    r["silver"],
                    r["bronze"]
                ])
        
        elif type == "grade":
            ws.title = "年级奖牌"
            headers = ["排名", "年级", "金牌", "银牌", "铜牌", "奖牌总数"]
            ws.append(headers)
            self._style_header(ws)
            
            rankings = self.stats_service.get_grade_medals()
            for r in rankings:
                ws.append([
                    r["rank"],
                    r["grade"]["name"],
                    r["gold"],
                    r["silver"],
                    r["bronze"],
                    r["total"]
                ])
        
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 12
        
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def export_participant_form(
        self,
        event_id: int,
        custom_fields: List[str] = None
    ) -> bytes:
        """导出参赛表格"""
        wb = self._create_workbook()
        ws = wb.active
        
        event = self.db.query(Event).filter(Event.id == event_id).first()
        ws.title = event.name if event else "参赛表格"
        
        # 默认字段
        default_fields = ["序号", "道次", "学号", "姓名", "班级", "年级"]
        headers = default_fields + (custom_fields or [])
        ws.append(headers)
        self._style_header(ws)
        
        registrations = self.db.query(Registration).filter(
            Registration.event_id == event_id
        ).order_by(Registration.lane_no).all()
        
        for idx, reg in enumerate(registrations, 1):
            row = [
                idx,
                reg.lane_no or idx,
                reg.student.student_no,
                reg.student.name,
                reg.student.class_.name,
                reg.student.class_.grade.name
            ]
            # 自定义字段留空
            row.extend([""] * len(custom_fields or []))
            ws.append(row)
        
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 12
        
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def export_all_events(self) -> bytes:
        """
        批量导出所有项目参赛表格（裁判成绩填写表）
        
        按项目+组别分工作表，包含成绩填写列和名次列
        """
        from app.models.event import EventGroup
        
        wb = self._create_workbook()
        wb.remove(wb.active)  # 移除默认sheet
        
        # 定义样式
        header_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        title_font = Font(bold=True, size=14)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        events = self.db.query(Event).order_by(Event.name).all()
        
        for event in events:
            # 获取该项目的组别
            groups = self.db.query(EventGroup).filter(EventGroup.event_id == event.id).all()
            
            # 如果没有组别，创建默认组
            if not groups:
                group_list = [{"id": None, "name": "默认组"}]
            else:
                group_list = [{"id": g.id, "name": g.name} for g in groups]
            
            for group_info in group_list:
                group_id = group_info["id"]
                group_name = group_info["name"]
                
                # 工作表名称：项目名-组别名（限制31字符）
                if group_name == "默认组":
                    sheet_name = event.name[:31]
                else:
                    sheet_name = f"{event.name}-{group_name}"[:31]
                
                # 避免重名
                base_name = sheet_name
                counter = 1
                while sheet_name in wb.sheetnames:
                    sheet_name = f"{base_name[:28]}_{counter}"
                    counter += 1
                
                ws = wb.create_sheet(title=sheet_name)
                
                # 添加标题行
                title_text = f"{event.name}"
                if group_name != "默认组":
                    title_text += f" - {group_name}"
                title_text += f"  （成绩单位：{event.unit}）"
                
                ws.cell(row=1, column=1, value=title_text)
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
                title_cell = ws.cell(row=1, column=1)
                title_cell.font = title_font
                title_cell.alignment = center_align
                
                # 表头
                headers = ["序号", "道次", "学号", "姓名", "性别", "班级", "年级", "成绩", "名次"]
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=3, column=col_idx, value=header)
                    cell.font = header_font
                    cell.alignment = center_align
                    cell.border = thin_border
                
                # 查询该项目-组别的报名记录
                reg_query = self.db.query(Registration).filter(
                    Registration.event_id == event.id
                )
                if group_id:
                    reg_query = reg_query.filter(Registration.group_id == group_id)
                else:
                    reg_query = reg_query.filter(Registration.group_id == None)
                
                registrations = reg_query.order_by(Registration.lane_no).all()
                
                # 数据行
                for idx, reg in enumerate(registrations, 1):
                    row_data = [
                        idx,
                        reg.lane_no or idx,
                        reg.student.student_no,
                        reg.student.name,
                        "男" if reg.student.gender == "M" else "女",
                        reg.student.class_.name,
                        reg.student.class_.grade.name,
                        "",  # 成绩（留空给裁判填写）
                        ""   # 名次（留空给裁判填写）
                    ]
                    
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=3 + idx, column=col_idx, value=value)
                        cell.alignment = center_align
                        cell.border = thin_border
                
                # 设置列宽
                column_widths = [8, 8, 15, 12, 8, 12, 12, 12, 8]
                for col_idx, width in enumerate(column_widths, 1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    # ========== 班级报名表导出功能 ==========
    
    def get_exportable_classes(self) -> List[Dict]:
        """
        获取有报名记录的班级列表
        
        Returns:
            包含班级信息的字典列表，每个字典包含:
            - id: 班级ID
            - name: 班级名称
            - grade_id: 年级ID
            - grade_name: 年级名称
            - registration_count: 报名人数
        """
        from sqlalchemy import func
        
        # 查询有报名记录的班级，统计每个班级的报名人数
        # 通过 Registration -> Student -> Class -> Grade 关联
        results = (
            self.db.query(
                Class.id,
                Class.name,
                Class.grade_id,
                Grade.name.label("grade_name"),
                func.count(Registration.id).label("registration_count")
            )
            .join(Student, Class.id == Student.class_id)
            .join(Registration, Student.id == Registration.student_id)
            .join(Grade, Class.grade_id == Grade.id)
            .group_by(Class.id, Class.name, Class.grade_id, Grade.name)
            .order_by(Grade.sort_order, Class.name)
            .all()
        )
        
        return [
            {
                "id": row.id,
                "name": row.name,
                "grade_id": row.grade_id,
                "grade_name": row.grade_name,
                "registration_count": row.registration_count
            }
            for row in results
        ]
    
    def export_class_registration_forms(self, class_ids: List[int]) -> bytes:
        """
        批量导出班级报名表
        
        Args:
            class_ids: 班级ID列表
            
        Returns:
            Excel文件的字节内容
            
        Raises:
            ValueError: 当class_ids为空或所有ID都无效时
            
        Requirements: 2.1, 2.2, 2.3
        """
        if not class_ids:
            raise ValueError("请至少选择一个班级")
        
        # 查询有效的班级信息
        valid_classes = (
            self.db.query(Class)
            .join(Grade, Class.grade_id == Grade.id)
            .filter(Class.id.in_(class_ids))
            .all()
        )
        
        if not valid_classes:
            raise ValueError("未找到有效班级")
        
        # 创建工作簿
        wb = self._create_workbook()
        wb.remove(wb.active)  # 移除默认sheet
        
        # 为每个班级创建工作表
        for class_ in valid_classes:
            # 工作表命名格式：年级名-班级名
            # Excel工作表名称限制31字符
            sheet_name = f"{class_.grade.name}-{class_.name}"[:31]
            ws = wb.create_sheet(title=sheet_name)
            
            # 构建工作表内容
            self._build_class_sheet(
                ws=ws,
                class_id=class_.id,
                class_name=class_.name,
                grade_name=class_.grade.name
            )
        
        # 保存到字节流
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def _group_registrations_by_event_group(
        self, 
        registrations: List[Registration]
    ) -> Dict[str, List[Registration]]:
        """
        按项目和组别分组报名记录
        
        Args:
            registrations: 报名记录列表
            
        Returns:
            分组后的字典，键为 "项目名称 - 组别名称" 格式，值为该组的报名记录列表
            无组别的报名记录归入 "项目名称 - 默认组"
            
        排序规则：
            - 先按项目名称排序
            - 再按组别名称排序
            - 组内按学号排序
        """
        from collections import defaultdict
        
        # 使用 defaultdict 收集分组
        groups: Dict[str, List[Registration]] = defaultdict(list)
        
        for reg in registrations:
            event_name = reg.event.name if reg.event else "未知项目"
            group_name = reg.group.name if reg.group else "默认组"
            key = f"{event_name} - {group_name}"
            groups[key].append(reg)
        
        # 对每个组内的记录按学号排序
        for key in groups:
            groups[key].sort(key=lambda r: r.student.student_no if r.student else "")
        
        # 按键（项目名称 - 组别名称）排序后返回
        sorted_groups = dict(sorted(groups.items()))
        
        return sorted_groups
    
    def _build_class_sheet(
        self, 
        ws: Worksheet, 
        class_id: int,
        class_name: str,
        grade_name: str
    ) -> None:
        """
        构建单个班级的工作表
        
        Args:
            ws: openpyxl工作表对象
            class_id: 班级ID
            class_name: 班级名称
            grade_name: 年级名称
            
        Requirements: 4.1, 4.2, 4.3, 4.4, 5.2, 5.3, 5.4
        """
        # 定义样式
        header_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        group_title_font = Font(bold=True, size=12)
        group_title_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 表头字段
        headers = ["序号", "学号", "姓名", "性别", "项目名称", "组别名称"]
        
        # 查询该班级的所有报名记录
        registrations = (
            self.db.query(Registration)
            .join(Student, Registration.student_id == Student.id)
            .filter(Student.class_id == class_id)
            .all()
        )
        
        # 按项目和组别分组
        grouped = self._group_registrations_by_event_group(registrations)
        
        current_row = 1
        
        # 遍历每个组别
        for group_key, group_regs in grouped.items():
            # 添加组别标题行（合并单元格）
            ws.cell(row=current_row, column=1, value=f"【{group_key}】")
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
            
            # 设置组别标题样式
            title_cell = ws.cell(row=current_row, column=1)
            title_cell.font = group_title_font
            title_cell.fill = group_title_fill
            title_cell.alignment = center_align
            
            current_row += 1
            
            # 添加表头行
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=header)
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
            
            current_row += 1
            
            # 添加数据行
            for seq_no, reg in enumerate(group_regs, 1):
                # 性别转换：M -> 男，F -> 女
                gender_display = "男" if reg.student.gender == "M" else "女"
                event_name = reg.event.name if reg.event else ""
                group_name = reg.group.name if reg.group else "默认组"
                
                row_data = [
                    seq_no,
                    reg.student.student_no,
                    reg.student.name,
                    gender_display,
                    event_name,
                    group_name
                ]
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    cell.alignment = center_align
                    cell.border = thin_border
                
                current_row += 1
            
            # 组别之间添加空行
            current_row += 1
        
        # 自动调整列宽
        column_widths = [8, 15, 12, 8, 15, 15]  # 预设列宽
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
