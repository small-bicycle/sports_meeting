"""
成绩管理服务模块
实现成绩录入、修改、作废、批量导入、查重
"""
from typing import List, Optional, Tuple, Dict
from decimal import Decimal
from sqlalchemy.orm import Session
from openpyxl import load_workbook
from io import BytesIO

from app.models.score import Score
from app.models.registration import Registration
from app.models.base import Student
from app.models.event import Event


class ScoreService:
    """成绩管理服务类"""
    
    def __init__(self, db: Session):
        self.db = db
    
    def check_duplicate(self, registration_id: int, round: str) -> Tuple[bool, Optional[Score]]:
        """
        检查是否重复录入成绩
        返回: (是否存在, 已存在的成绩记录)
        """
        existing = self.db.query(Score).filter(
            Score.registration_id == registration_id,
            Score.round == round,
            Score.is_valid == True
        ).first()
        return existing is not None, existing
    
    def create_score(
        self,
        registration_id: int,
        value: float,
        round: str = "final",
        created_by: int = None,
        overwrite: bool = False
    ) -> Tuple[Optional[Score], str]:
        """
        录入成绩
        overwrite: 是否覆盖已存在的成绩
        """
        # 检查报名记录
        registration = self.db.query(Registration).filter(Registration.id == registration_id).first()
        if not registration:
            return None, "报名记录不存在"
        
        if round not in ("preliminary", "final"):
            return None, "轮次必须是preliminary(预赛)或final(决赛)"
        
        # 检查重复
        is_duplicate, existing = self.check_duplicate(registration_id, round)
        if is_duplicate:
            if not overwrite:
                return None, "该轮次已有成绩记录，是否覆盖？"
            # 作废旧成绩
            existing.is_valid = False
            existing.invalid_reason = "被新成绩覆盖"
        
        score = Score(
            registration_id=registration_id,
            value=Decimal(str(value)),
            round=round,
            created_by=created_by
        )
        self.db.add(score)
        self.db.commit()
        self.db.refresh(score)
        return score, ""
    
    def update_score(
        self,
        score_id: int,
        value: float,
        reason: str,
        updated_by: int = None
    ) -> Tuple[Optional[Score], str]:
        """修改成绩，保留修改记录"""
        score = self.db.query(Score).filter(Score.id == score_id).first()
        if not score:
            return None, "成绩记录不存在"
        
        if not score.is_valid:
            return None, "该成绩已作废，不能修改"
        
        if not reason:
            return None, "请填写修改原因"
        
        # 记录修改
        score.value = Decimal(str(value))
        score.update_reason = reason
        score.updated_by = updated_by
        
        self.db.commit()
        self.db.refresh(score)
        return score, ""
    
    def invalidate_score(
        self,
        score_id: int,
        reason: str,
        updated_by: int = None
    ) -> Tuple[Optional[Score], str]:
        """作废成绩，保留原始数据"""
        score = self.db.query(Score).filter(Score.id == score_id).first()
        if not score:
            return None, "成绩记录不存在"
        
        if not score.is_valid:
            return None, "该成绩已作废"
        
        if not reason:
            return None, "请填写作废原因"
        
        score.is_valid = False
        score.invalid_reason = reason
        score.updated_by = updated_by
        
        self.db.commit()
        self.db.refresh(score)
        return score, ""
    
    def get_score_list(
        self,
        page: int = 1,
        page_size: int = 20,
        event_id: int = None,
        class_id: int = None,
        grade_id: int = None,
        student_id: int = None,
        round: str = None,
        include_invalid: bool = False
    ) -> Tuple[List[Score], int]:
        """获取成绩列表"""
        query = self.db.query(Score).join(Registration).join(Student)
        
        if not include_invalid:
            query = query.filter(Score.is_valid == True)
        
        if event_id:
            query = query.filter(Registration.event_id == event_id)
        if class_id:
            query = query.filter(Student.class_id == class_id)
        if grade_id:
            from app.models.base import Class
            query = query.join(Class).filter(Class.grade_id == grade_id)
        if student_id:
            query = query.filter(Registration.student_id == student_id)
        if round:
            query = query.filter(Score.round == round)
        
        total = query.count()
        scores = query.offset((page - 1) * page_size).limit(page_size).all()
        
        return scores, total
    
    def get_score_by_id(self, score_id: int) -> Optional[Score]:
        """根据ID获取成绩"""
        return self.db.query(Score).filter(Score.id == score_id).first()
    
    def import_scores(
        self, 
        file_content: bytes, 
        round: str = "final",
        created_by: int = None
    ) -> Tuple[int, int, List[str]]:
        """
        批量导入成绩
        Excel格式: 学号, 项目名称, 成绩
        返回: (成功数, 失败数, 错误列表)
        """
        success_count = 0
        fail_count = 0
        errors = []
        
        try:
            wb = load_workbook(BytesIO(file_content))
            ws = wb.active
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:
                    continue
                
                try:
                    student_no = str(row[0]).strip()
                    event_name = str(row[1]).strip()
                    value = float(row[2])
                    
                    # 查找学生
                    student = self.db.query(Student).filter(Student.student_no == student_no).first()
                    if not student:
                        errors.append(f"第{row_idx}行: 学号'{student_no}'不存在")
                        fail_count += 1
                        continue
                    
                    # 查找项目
                    event = self.db.query(Event).filter(Event.name == event_name).first()
                    if not event:
                        errors.append(f"第{row_idx}行: 项目'{event_name}'不存在")
                        fail_count += 1
                        continue
                    
                    # 查找报名记录
                    registration = self.db.query(Registration).filter(
                        Registration.student_id == student.id,
                        Registration.event_id == event.id
                    ).first()
                    if not registration:
                        errors.append(f"第{row_idx}行: 该学生未报名此项目")
                        fail_count += 1
                        continue
                    
                    # 录入成绩
                    score, error = self.create_score(
                        registration_id=registration.id,
                        value=value,
                        round=round,
                        created_by=created_by,
                        overwrite=True
                    )
                    
                    if error and "是否覆盖" not in error:
                        errors.append(f"第{row_idx}行: {error}")
                        fail_count += 1
                    else:
                        success_count += 1
                        
                except ValueError:
                    errors.append(f"第{row_idx}行: 成绩格式错误")
                    fail_count += 1
                except Exception as e:
                    errors.append(f"第{row_idx}行: {str(e)}")
                    fail_count += 1
            
            self.db.commit()
            
        except Exception as e:
            errors.append(f"文件解析错误: {str(e)}")
        
        return success_count, fail_count, errors
